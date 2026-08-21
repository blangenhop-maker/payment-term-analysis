#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Payment Term Analysis
---------------------
Streamlit app to analyze contract payment terms versus supplier payment terms.

Primary use case:
1. Upload an Excel or CSV Analytics export.
2. Identify vendors in column B that appear more than once.
3. Compare column G, Contract Payment Term, against column I, Supplier Payment Term.
4. Flag mismatches.
5. Return all mismatch rows.
6. Also return all additional contracts for vendors that have at least one mismatch.

Run with:
    streamlit run PaymentTermAnalysis.py

Expected default column positions, if matching headers are not found:
    Column B = Supplier / Vendor name
    Column G = Contract Payment Term
    Column I = Supplier Payment Term
"""

from __future__ import annotations

import io
import re
import warnings
from contextlib import contextmanager
from datetime import datetime
from typing import Dict, Optional, Tuple, List

import pandas as pd
import streamlit as st


# --------------------- Warning control ---------------------

@contextmanager
def ignore_openpyxl_default_style_warning():
    """Temporarily silence openpyxl's 'Workbook contains no default style' UserWarning."""
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style",
            category=UserWarning,
            module="openpyxl",
        )
        yield


# --------------------- Constants ---------------------

DEFAULT_VENDOR_COL_INDEX = 1      # Column B
DEFAULT_CONTRACT_TERM_INDEX = 6   # Column G
DEFAULT_SUPPLIER_TERM_INDEX = 8   # Column I

VENDOR_COLUMN_CANDIDATES = [
    "Supplier Display Name",
    "Supplier Name",
    "Vendor Name",
    "Vendor",
    "Supplier",
]

CONTRACT_TERM_COLUMN_CANDIDATES = [
    "Contract Payment Term",
    "Contract Payment Terms",
    "Contract Term",
    "Payment Term",
]

SUPPLIER_TERM_COLUMN_CANDIDATES = [
    "Supplier Payment Term",
    "Supplier Payment Terms",
    "Supplier Term",
    "Vendor Payment Term",
    "Vendor Payment Terms",
]

CONTRACT_TAG_COLUMN_CANDIDATES = [
    "Contract Tags",
    "Contract Tag",
    "Tags",
    "Tag",
    "Contract Labels",
    "Contract Label",
]


# --------------------- Utility functions ---------------------

def normalize_header(value) -> str:
    """Normalize a column header for loose matching."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def find_column_by_candidates(df: pd.DataFrame, candidates: List[str], fallback_index: int) -> str:
    """
    Find a DataFrame column by a list of likely header names.
    If no header match is found, use the fallback Excel-style zero-based index.
    """
    normalized_map = {normalize_header(col): col for col in df.columns}

    for candidate in candidates:
        key = normalize_header(candidate)
        if key in normalized_map:
            return normalized_map[key]

    if fallback_index >= len(df.columns):
        raise ValueError(
            f"The file does not contain enough columns for fallback column index {fallback_index}. "
            f"Detected only {len(df.columns)} columns."
        )

    return df.columns[fallback_index]


def find_optional_column_by_candidates(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """Find a DataFrame column by likely header names, returning None if no match exists."""
    normalized_map = {normalize_header(col): col for col in df.columns}

    for candidate in candidates:
        key = normalize_header(candidate)
        if key in normalized_map:
            return normalized_map[key]

    return None


def normalize_tag_text(value) -> str:
    """Normalize a tag field for case-insensitive tag checks."""
    return re.sub(r"\s+", " ", clean_text(value)).casefold()


def has_dual_payment_terms_tag(value) -> bool:
    """Return True when the contract tags include Dual Payment Terms."""
    return "dual payment terms" in normalize_tag_text(value)


def clean_text(value) -> str:
    """Convert a cell value to a trimmed string while preserving blanks."""
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_payment_term(value) -> str:
    """
    Normalize payment terms to reduce false mismatches.

    Examples treated as equivalent:
        Net 30
        NET30
        Net-30
        Net 30 Days
    """
    text = clean_text(value).upper()

    # Standardize common punctuation and wording.
    text = text.replace("&", " AND ")
    text = text.replace("-", " ")
    text = text.replace("_", " ")
    text = text.replace("/", " ")
    text = re.sub(r"\bDAYS\b", "", text)
    text = re.sub(r"\bDAY\b", "", text)

    # Remove all non-alphanumeric characters for comparison.
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


def read_excel_sheet_names(file_like) -> List[str]:
    """Return sheet names from an uploaded Excel file-like object."""
    file_like.seek(0)
    with ignore_openpyxl_default_style_warning():
        xls = pd.ExcelFile(file_like, engine="openpyxl")
    return xls.sheet_names


def read_uploaded_table(uploaded_file, chosen_sheet: Optional[str] = None) -> pd.DataFrame:
    """
    Read one table from an uploaded CSV or Excel file.
    All fields are read as text to preserve leading zeros and exact values.
    """
    uploaded_file.seek(0)
    file_name = uploaded_file.name.lower()

    if file_name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
    else:
        with ignore_openpyxl_default_style_warning():
            df = pd.read_excel(
                uploaded_file,
                sheet_name=chosen_sheet if chosen_sheet else 0,
                dtype=str,
                keep_default_na=False,
                engine="openpyxl",
            )

    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    return df


def autosize_worksheet_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Set reasonable Excel column widths for readability."""
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, start=1):
        values = df[col].astype(str).head(500).tolist() if not df.empty else []
        max_len = max([len(str(col))] + [len(v) for v in values])
        width = min(max(max_len + 2, 12), 50)
        worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = width


def apply_basic_formatting(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Freeze header row, apply filters, and optionally highlight mismatch rows."""
    from openpyxl.styles import Font, PatternFill

    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"

    if not df.empty:
        worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    include_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    if "Term Mismatch" in df.columns:
        mismatch_col = list(df.columns).index("Term Mismatch") + 1
        for row_idx in range(2, len(df) + 2):
            if str(worksheet.cell(row=row_idx, column=mismatch_col).value).upper() == "YES":
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = mismatch_fill

    if "Included Because Vendor Has Mismatch" in df.columns:
        include_col = list(df.columns).index("Included Because Vendor Has Mismatch") + 1
        for row_idx in range(2, len(df) + 2):
            include_val = str(worksheet.cell(row=row_idx, column=include_col).value).upper()
            if include_val == "YES":
                # Only use the yellow fill when the row was not already marked as an actual mismatch.
                already_red = False
                if "Term Mismatch" in df.columns:
                    mismatch_col = list(df.columns).index("Term Mismatch") + 1
                    already_red = str(worksheet.cell(row=row_idx, column=mismatch_col).value).upper() == "YES"
                if not already_red:
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = include_fill

    autosize_worksheet_columns(writer, sheet_name, df)


def build_analysis(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, int], Dict[str, str]]:
    """
    Build all analysis tabs.

    Returns:
        analyzed_df
        summary_df
        vendor_summary_df
        mismatches_only_df
        vendor_contract_review_df
        duplicate_vendors_df
        metrics
        selected_columns
    """
    if df.empty:
        raise ValueError("The uploaded file did not contain any rows.")

    vendor_col = find_column_by_candidates(df, VENDOR_COLUMN_CANDIDATES, DEFAULT_VENDOR_COL_INDEX)
    contract_term_col = find_column_by_candidates(df, CONTRACT_TERM_COLUMN_CANDIDATES, DEFAULT_CONTRACT_TERM_INDEX)
    supplier_term_col = find_column_by_candidates(df, SUPPLIER_TERM_COLUMN_CANDIDATES, DEFAULT_SUPPLIER_TERM_INDEX)
    contract_tag_col = find_optional_column_by_candidates(df, CONTRACT_TAG_COLUMN_CANDIDATES)

    analyzed = df.copy()

    analyzed["Vendor Name - Clean"] = analyzed[vendor_col].apply(clean_text)
    analyzed["Contract Payment Term - Clean"] = analyzed[contract_term_col].apply(clean_text)
    analyzed["Supplier Payment Term - Clean"] = analyzed[supplier_term_col].apply(clean_text)
    analyzed["Contract Payment Term - Normalized"] = analyzed[contract_term_col].apply(normalize_payment_term)
    analyzed["Supplier Payment Term - Normalized"] = analyzed[supplier_term_col].apply(normalize_payment_term)

    if contract_tag_col:
        analyzed["Contract Tags - Clean"] = analyzed[contract_tag_col].apply(clean_text)
        analyzed["Has Dual Payment Terms Tag"] = analyzed[contract_tag_col].apply(
            lambda value: "YES" if has_dual_payment_terms_tag(value) else "NO"
        )
    else:
        analyzed["Contract Tags - Clean"] = ""
        analyzed["Has Dual Payment Terms Tag"] = "NO"

    # Vendor counts identify duplicate vendors, meaning vendors with multiple contracts/rows.
    vendor_counts = analyzed["Vendor Name - Clean"].value_counts(dropna=False)
    analyzed["Contract Count for Vendor"] = analyzed["Vendor Name - Clean"].map(vendor_counts).fillna(0).astype(int)
    analyzed["Multiple Contracts"] = analyzed["Contract Count for Vendor"].apply(lambda x: "YES" if x > 1 else "NO")

    # Flag mismatches using normalized terms.
    analyzed["Term Mismatch"] = analyzed.apply(
        lambda row: "YES"
        if row["Contract Payment Term - Normalized"] != row["Supplier Payment Term - Normalized"]
        else "NO",
        axis=1,
    )

    mismatch_vendor_set = set(
        analyzed.loc[analyzed["Term Mismatch"] == "YES", "Vendor Name - Clean"].dropna().tolist()
    )

    analyzed["Included Because Vendor Has Mismatch"] = analyzed.apply(
        lambda row: "YES"
        if row["Vendor Name - Clean"] in mismatch_vendor_set and row["Term Mismatch"] != "YES"
        else "NO",
        axis=1,
    )

    analyzed["Analysis Result"] = analyzed.apply(
        lambda row: "Actual term mismatch"
        if row["Term Mismatch"] == "YES"
        else (
            "Additional contract for mismatched vendor"
            if row["Included Because Vendor Has Mismatch"] == "YES"
            else "No mismatch"
        ),
        axis=1,
    )

    vendor_dual_tag_counts = analyzed.groupby("Vendor Name - Clean", dropna=False)[
        "Has Dual Payment Terms Tag"
    ].apply(lambda s: int((s == "YES").sum()))
    analyzed["Dual Payment Terms Tag Count for Vendor"] = analyzed["Vendor Name - Clean"].map(
        vendor_dual_tag_counts
    ).fillna(0).astype(int)
    analyzed["All Vendor Contracts Tagged Dual Payment Terms"] = analyzed.apply(
        lambda row: "YES"
        if row["Contract Count for Vendor"] > 0
        and row["Dual Payment Terms Tag Count for Vendor"] == row["Contract Count for Vendor"]
        else "NO",
        axis=1,
    )

    def review_comment(row) -> str:
        """Create a review comment that combines mismatch status, contract count, and tags."""
        count = int(row["Contract Count for Vendor"])
        has_dual = row["Has Dual Payment Terms Tag"] == "YES"
        all_dual = row["All Vendor Contracts Tagged Dual Payment Terms"] == "YES"
        mismatch = row["Term Mismatch"] == "YES"
        included_context = row["Included Because Vendor Has Mismatch"] == "YES"

        contract_word = "contract" if count == 1 else "contracts"
        tag_count = int(row["Dual Payment Terms Tag Count for Vendor"])
        tag_word = "contract" if tag_count == 1 else "contracts"

        if count == 1 and has_dual:
            return "Review needed: contract is tagged Dual Payment Terms, but vendor has 1 contract in Coupa."

        if count > 1 and all_dual and (mismatch or included_context):
            return f"Likely OK to ignore: vendor has {count} {contract_word} in Coupa and all contracts are tagged Dual Payment Terms."

        if mismatch and has_dual and count > 1:
            return f"Review needed: mismatch is tagged Dual Payment Terms, but only {tag_count} of {count} vendor contracts in Coupa are tagged Dual Payment Terms."

        if mismatch and not has_dual:
            return f"Review needed: payment-term mismatch is not covered by a Dual Payment Terms tag. Vendor has {count} {contract_word} in Coupa."

        if included_context and has_dual:
            return f"Context only: matching contract for a mismatched vendor. This contract is tagged Dual Payment Terms, and vendor has {count} {contract_word} in Coupa."

        if included_context:
            return f"Context only: matching contract included because this vendor has another payment-term mismatch. Vendor has {count} {contract_word} in Coupa."

        if has_dual:
            return f"Review tag: Dual Payment Terms tag exists, but no payment-term mismatch was identified on this row. Vendor has {count} {contract_word} in Coupa."

        return "No issue identified by current rules."

    analyzed["Review Comment"] = analyzed.apply(review_comment, axis=1)

    duplicate_vendors = (
        analyzed.loc[analyzed["Contract Count for Vendor"] > 1]
        .groupby("Vendor Name - Clean", dropna=False)
        .agg(
            Contract_Count=("Vendor Name - Clean", "size"),
            Mismatch_Count=("Term Mismatch", lambda s: int((s == "YES").sum())),
        )
        .reset_index()
        .rename(columns={"Vendor Name - Clean": "Vendor Name"})
        .sort_values(["Mismatch_Count", "Contract_Count", "Vendor Name"], ascending=[False, False, True])
    )

    vendor_summary = (
        analyzed.groupby("Vendor Name - Clean", dropna=False)
        .agg(
            Contract_Count=("Vendor Name - Clean", "size"),
            Mismatch_Count=("Term Mismatch", lambda s: int((s == "YES").sum())),
            Multiple_Contracts=("Multiple Contracts", lambda s: "YES" if (s == "YES").any() else "NO"),
            Included_In_Review=("Vendor Name - Clean", lambda s: "YES" if s.iloc[0] in mismatch_vendor_set else "NO"),
        )
        .reset_index()
        .rename(columns={"Vendor Name - Clean": "Vendor Name"})
        .sort_values(["Mismatch_Count", "Contract_Count", "Vendor Name"], ascending=[False, False, True])
    )

    mismatches_only = analyzed.loc[analyzed["Term Mismatch"] == "YES"].copy()

    vendor_contract_review = analyzed.loc[
        analyzed["Vendor Name - Clean"].isin(mismatch_vendor_set)
    ].copy()

    sort_cols = ["Vendor Name - Clean", "Term Mismatch"]
    vendor_contract_review = vendor_contract_review.sort_values(sort_cols, ascending=[True, False])
    mismatches_only = mismatches_only.sort_values(["Vendor Name - Clean"])

    total_contracts = len(analyzed)
    vendors_total = analyzed["Vendor Name - Clean"].nunique(dropna=False)
    vendors_with_multiple = int((vendor_counts > 1).sum())
    mismatched_contracts = int((analyzed["Term Mismatch"] == "YES").sum())
    vendors_with_mismatches = len(mismatch_vendor_set)
    extra_contracts_for_mismatched_vendors = int(
        (vendor_contract_review["Term Mismatch"] != "YES").sum()
    )

    metrics = {
        "Total contracts / rows": total_contracts,
        "Total vendors": vendors_total,
        "Vendors with multiple contracts": vendors_with_multiple,
        "Mismatched contracts": mismatched_contracts,
        "Vendors with at least one mismatch": vendors_with_mismatches,
        "Additional non-mismatch contracts included for mismatched vendors": extra_contracts_for_mismatched_vendors,
    }

    selected_columns = {
        "Vendor column used": str(vendor_col),
        "Contract payment term column used": str(contract_term_col),
        "Supplier payment term column used": str(supplier_term_col),
        "Contract tags column used": str(contract_tag_col) if contract_tag_col else "No contract tag column detected",
    }

    summary_rows = []
    for metric, count in metrics.items():
        summary_rows.append({"Metric": metric, "Value": count})
    for label, value in selected_columns.items():
        summary_rows.append({"Metric": label, "Value": value})

    summary = pd.DataFrame(summary_rows)

    # Put the key analysis columns first while preserving all source columns.
    priority_cols = [
        "Vendor Name - Clean",
        "Contract Payment Term - Clean",
        "Supplier Payment Term - Clean",
        "Term Mismatch",
        "Multiple Contracts",
        "Contract Count for Vendor",
        "Included Because Vendor Has Mismatch",
        "Analysis Result",
        "Review Comment",
        "Contract Tags - Clean",
        "Has Dual Payment Terms Tag",
        "Dual Payment Terms Tag Count for Vendor",
        "All Vendor Contracts Tagged Dual Payment Terms",
        "Contract Payment Term - Normalized",
        "Supplier Payment Term - Normalized",
    ]
    source_cols = [c for c in analyzed.columns if c not in priority_cols]
    analyzed = analyzed[priority_cols + source_cols]
    mismatches_only = mismatches_only[[c for c in analyzed.columns if c in mismatches_only.columns]]
    vendor_contract_review = vendor_contract_review[[c for c in analyzed.columns if c in vendor_contract_review.columns]]

    return (
        analyzed,
        summary,
        vendor_summary,
        mismatches_only,
        vendor_contract_review,
        duplicate_vendors,
        metrics,
        selected_columns,
    )


def write_output_workbook(
    analyzed: pd.DataFrame,
    summary: pd.DataFrame,
    vendor_summary: pd.DataFrame,
    mismatches_only: pd.DataFrame,
    vendor_contract_review: pd.DataFrame,
    duplicate_vendors: pd.DataFrame,
) -> bytes:
    """Write the analysis output to an in-memory Excel workbook."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        vendor_summary.to_excel(writer, index=False, sheet_name="Vendor Summary")
        mismatches_only.to_excel(writer, index=False, sheet_name="Mismatches Only")
        vendor_contract_review.to_excel(writer, index=False, sheet_name="Vendor Contract Review")
        duplicate_vendors.to_excel(writer, index=False, sheet_name="Duplicate Vendors")
        analyzed.to_excel(writer, index=False, sheet_name="All Rows Analyzed")

        for sheet_name, df in {
            "Summary": summary,
            "Vendor Summary": vendor_summary,
            "Mismatches Only": mismatches_only,
            "Vendor Contract Review": vendor_contract_review,
            "Duplicate Vendors": duplicate_vendors,
            "All Rows Analyzed": analyzed,
        }.items():
            apply_basic_formatting(writer, sheet_name, df)

    output.seek(0)
    return output.getvalue()


# --------------------- Streamlit UI ---------------------

def main() -> None:
    st.set_page_config(page_title="Payment Term Analysis", page_icon="https://arc.workspaceoneaccess.com/catalog-portal/services/api/resources/media/b7bbd3c4-ade2-4d72-a9c9-006abb088f59", layout="wide")

    st.title("Contract vs Supplier Payment Term Analysis")
    st.caption(
        "Upload a Coupa Analytics export and identify vendors with multiple contracts, payment-term mismatches, "
        "and all additional contracts for vendors that have at least one mismatch."
    )

    with st.sidebar:
        st.header("Analysis Logic")
        st.markdown(
            """
            **Default columns used if headers are not found:**
            - Column B: Vendor / Supplier name
            - Column G: Contract payment term
            - Column I: Supplier payment term

            **Output includes:**
            - Summary
            - Vendor Summary
            - Mismatches Only
            - Vendor Contract Review
            - Duplicate Vendors
            - All Rows Analyzed
            """
        )

    uploaded_file = st.file_uploader(
        "Upload payment term analysis file (.xlsx or .csv)",
        type=["xlsx", "csv"],
        accept_multiple_files=False,
    )

    chosen_sheet = None
    if uploaded_file and uploaded_file.name.lower().endswith(".xlsx"):
        try:
            sheet_names = read_excel_sheet_names(uploaded_file)
            chosen_sheet = st.selectbox(
                "Select worksheet",
                options=sheet_names,
                index=0,
                help="Choose the worksheet that contains the vendor, contract term, and supplier term columns.",
            )
        except Exception as exc:
            st.error(f"Could not read worksheet names: {exc}")
            st.stop()

    if not uploaded_file:
        st.info("Upload an Excel or CSV file to begin.")
        return

    try:
        df = read_uploaded_table(uploaded_file, chosen_sheet)
    except Exception as exc:
        st.error(f"Could not read the uploaded file: {exc}")
        st.stop()

    st.subheader("Preview of Uploaded Data")
    st.dataframe(df.head(25), width="stretch")

    build = st.button("⚙️ Build Analysis Workbook", type="primary")

    if build:
        try:
            (
                analyzed,
                summary,
                vendor_summary,
                mismatches_only,
                vendor_contract_review,
                duplicate_vendors,
                metrics,
                selected_columns,
            ) = build_analysis(df)

            output_bytes = write_output_workbook(
                analyzed=analyzed,
                summary=summary,
                vendor_summary=vendor_summary,
                mismatches_only=mismatches_only,
                vendor_contract_review=vendor_contract_review,
                duplicate_vendors=duplicate_vendors,
            )

            st.success("Analysis complete.")

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Contracts", metrics["Total contracts / rows"])
            col2.metric("Duplicate Vendors", metrics["Vendors with multiple contracts"])
            col3.metric("Mismatched Contracts", metrics["Mismatched contracts"])
            col4.metric("Vendors with Mismatches", metrics["Vendors with at least one mismatch"])

            st.subheader("Columns Used")
            st.write(selected_columns)

            st.subheader("Mismatches Only")
            st.dataframe(mismatches_only, width="stretch")

            st.subheader("Vendor Contract Review")
            st.caption(
                "This tab includes every contract for vendors that have at least one payment-term mismatch. "
                "Rows marked as 'Additional contract for mismatched vendor' are included for context. "
                "The Review Comment column uses contract tags and contract count to suggest whether the row is likely OK or needs review."
            )
            st.dataframe(vendor_contract_review, width="stretch")

            date_stamp = datetime.now().strftime("%m%d%Y")
            output_name = f"Payment_Term_Analysis_{date_stamp}.xlsx"

            st.download_button(
                label="⬇️ Download Analysis Workbook",
                data=output_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as exc:
            st.error(f"Analysis failed: {exc}")
            st.stop()


if __name__ == "__main__":
    main()
